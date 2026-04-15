# -*- coding: utf-8 -*-
import openpyxl
import json
import os
from datetime import datetime

print("🚀 Generando datos actualizados...")

ruta_datos = "datos_privados/"

# 1. CUMPLEAÑOS
wb = openpyxl.load_workbook(ruta_datos + "Alumnos.xlsx")
ws = wb["export-2"]
eventos = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or len(row) < 4: continue
    nombre = str(row[0]).strip() if row[0] else ""
    turma = str(row[1]).strip() if row[1] else ""
    fecha_raw = row[3]
    if nombre and fecha_raw:
        try:
            fecha = fecha_raw.strftime('%d/%m')
        except:
            fecha = str(fecha_raw)[:5]
        eventos.append({
            "fecha": fecha,
            "tipo": "cumpleaños",
            "descripcion": "Aniversário de {} {}".format(nombre, turma)
        })
print(" • {} cumpleaños cargados".format(len(eventos)))

# 2. DÍAS ESPECIALES + MENÚ
wb_menu = openpyxl.load_workbook(ruta_datos + "Menu del dia.xlsx")

# Días especiales
ws_especiales = wb_menu["DiasEspeciales"]
for row in ws_especiales.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    fecha_raw = row[0]
    desc = str(row[1]).strip() if len(row) > 1 else ""
    if fecha_raw and desc and desc != "nan" and desc != "":
        try:
            fecha = fecha_raw.strftime('%d/%m')
        except:
            fecha = str(fecha_raw)[:5]
        eventos.append({"fecha": fecha, "tipo": "especial", "descripcion": desc})
print(" • {} días especiales cargados".format(len([e for e in eventos if e['tipo'] == 'especial'])))

# Menú
ws_menu = wb_menu["MenuSimple"]
menu = {}
for row in ws_menu.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    fecha_raw = row[0]
    desc = str(row[1]).strip() if len(row) > 1 else ""
    if desc and len(desc) > 10:
        try:
            fecha = fecha_raw.strftime('%d/%m')
        except:
            fecha = str(fecha_raw)[:5]
        menu[fecha] = desc
print(" • {} días de menú cargados".format(len(menu)))

# 3. FOTOS Y VÍDEOS
imagenes = ["fotos/" + f for f in os.listdir("fotos") if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.mp4', '.webm', '.mov', '.avi'))]
print(" • {} archivos en galería (fotos + vídeos)".format(len(imagenes)))

# 4. MENSAJE PERSONALIZADO
mensaje_hoy = ""
try:
    ws_mensaje = wb_menu["MensajeDia"]
    hoy_str = datetime.today().strftime('%d/%m')
    for row in ws_mensaje.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        fecha_raw = row[0]
        texto = str(row[1]).strip() if len(row) > 1 else ""
        try:
            fecha = fecha_raw.strftime('%d/%m')
        except:
            fecha = str(fecha_raw)[:5]
        if fecha == hoy_str and texto and texto != "nan":
            mensaje_hoy = texto
            break
    with open("mensaje_dia.json", "w", encoding="utf-8") as f:
        json.dump(mensaje_hoy, f, ensure_ascii=False)
    print(" • Mensaje personalizado cargado" if mensaje_hoy else " • No hay mensaje personalizado hoy → se usa Wikipedia")
except:
    print(" • No existe hoja MensajeDia o hay un error")

# GUARDAR JSONs
with open("eventos.json", "w", encoding="utf-8") as f:
    json.dump(eventos, f, ensure_ascii=False, indent=2)
with open("menu.json", "w", encoding="utf-8") as f:
    json.dump(menu, f, ensure_ascii=False, indent=2)
with open("fotos.json", "w", encoding="utf-8") as f:
    json.dump(imagenes, f, ensure_ascii=False, indent=2)

print("✅ TODO GENERADO CORRECTAMENTE")